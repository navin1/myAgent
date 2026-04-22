"""
Mapping service: semantic parsing of Excel transformation specification files.

Classification is folder-based (not filename-based):
  - Mapping Sheets : files inside EXCEL_MAPPING_PATH (and its subfolders)
                     multi-row header (EXCEL_MAPPING_HEADER_ROWS), source→target column logic
  - Master Sheets  : every other Excel file under EXCEL_DATA_PATH
                     single or multi-row header (EXCEL_MASTER_HEADER_ROWS), structural definitions

The resulting index lets the agent answer:
  "What columns does table X map from source to target?"
  "What transformation logic is defined for column Y?"
"""
import logging
import re
from pathlib import Path
from typing import Any

import pandas as pd

import settings

log = logging.getLogger(__name__)

# Normalised-name → parsed spec
_index: dict[str, dict[str, Any]] = {}
_loaded: bool = False


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _normalize_name(name: str) -> str:
    """Any naming convention → lowercase_snake_case for stable lookups."""
    s = re.sub(r"(.)([A-Z][a-z]+)", r"\1_\2", name)
    s = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", s)
    return re.sub(r"[^a-z0-9]+", "_", s.lower()).strip("_")


def _classify_file(path: Path, mapping_root: Path | None) -> str:
    """
    Return 'mapping' if *path* lives inside *mapping_root*, 'master' otherwise.
    Classification is purely folder-based — filenames are irrelevant.
    """
    if mapping_root and path.is_relative_to(mapping_root):
        return "mapping"
    return "master"


def _flatten_multiindex(df: pd.DataFrame) -> pd.DataFrame:
    """
    Collapse a MultiIndex column header into single-level strings.
    Levels are joined with " | " so origin is still readable.
    """
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [
            " | ".join(
                str(c).strip() for c in col
                if str(c).strip() not in ("", "nan", "None")
            )
            for col in df.columns
        ]
    else:
        df.columns = [str(c).strip() for c in df.columns]
    return df


def _extract_source_target_pairs(df: pd.DataFrame) -> list[dict[str, str]]:
    """
    Heuristically locate source-column, target-column, and transformation columns
    from a flattened mapping sheet and return structured pairs.

    Detection logic (first match wins):
      - source col: column name contains both "source" (or "src") and "column" (or "col" or "field")
      - target col: column name contains both "target" (or "tgt" or "dest") and "column" (or "col" or "field")
      - transformation: column name contains any of "transform", "logic", "rule", "expression", "formula", "derivation"
    """
    cols = list(df.columns)

    def _find(keywords_a: list[str], keywords_b: list[str]) -> str | None:
        for c in cols:
            cl = c.lower()
            if any(k in cl for k in keywords_a) and any(k in cl for k in keywords_b):
                return c
        return None

    src_col = _find(["source", "src"], ["column", "col", "field", "attribute"])
    tgt_col = _find(["target", "tgt", "destination", "dest"], ["column", "col", "field", "attribute"])
    tfm_col = _find(
        ["transform", "logic", "rule", "expression", "formula", "derivation", "calculation", "mapping"],
        [""],  # any column containing these words qualifies
    )
    # Special case: single keyword match for transformation
    if not tfm_col:
        tfm_col = next(
            (c for c in cols if any(
                k in c.lower()
                for k in ("transform", "logic", "rule", "expression", "formula", "derivation")
            )),
            None,
        )

    if not src_col and not tgt_col:
        return []

    pairs: list[dict[str, str]] = []
    for _, row in df.iterrows():
        entry: dict[str, str] = {}
        if src_col and pd.notna(row.get(src_col)):
            v = str(row[src_col]).strip()
            if v:
                entry["source_column"] = v
        if tgt_col and pd.notna(row.get(tgt_col)):
            v = str(row[tgt_col]).strip()
            if v:
                entry["target_column"] = v
        if tfm_col and pd.notna(row.get(tfm_col)):
            v = str(row[tfm_col]).strip()
            if v and v.lower() not in ("nan", "none", "-"):
                entry["transformation"] = v
        if entry:
            pairs.append(entry)

    return pairs


def _parse_sheet(
    df: pd.DataFrame,
    file_type: str,
) -> dict[str, Any]:
    """Return a parsed representation of a single sheet."""
    df = _flatten_multiindex(df)

    # Trim rows that are completely empty
    df = df.dropna(how="all").reset_index(drop=True)

    sheet_info: dict[str, Any] = {
        "columns": list(df.columns),
        "row_count": len(df),
    }

    if file_type == "mapping":
        pairs = _extract_source_target_pairs(df)
        if pairs:
            sheet_info["source_target_pairs"] = pairs
            sheet_info["mapped_columns"] = len(pairs)

    # Store raw data (capped for LLM context; full data available via DuckDB query)
    records = df.where(df.notna(), None).to_dict(orient="records")
    sheet_info["sample_rows"] = records[:50]
    if len(records) > 50:
        sheet_info["sample_note"] = f"{len(records)} rows total; showing first 50. Use query_excel() for full data."

    return sheet_info


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------

def _header_arg(file_type: str) -> int | list[int]:
    """Return the pandas header= argument for a given file type.

    Mapping files: EXCEL_MAPPING_HEADER_ROWS is a 1-indexed row offset —
    rows before it are skipped and that row becomes the single column header.

    Master files: EXCEL_MASTER_HEADER_ROWS rows form a MultiIndex header.
    """
    if file_type == "mapping":
        return settings.EXCEL_MAPPING_HEADER_ROWS - 1  # 0-indexed header row
    n = settings.EXCEL_MASTER_HEADER_ROWS
    return list(range(n)) if n > 1 else 0


def _read_bq_table_name(file_path: Path) -> str | None:
    """Read raw [Row 1, Col 1] as the BigQuery target table name before header rows are skipped."""
    try:
        import pandas as pd
        raw = pd.read_excel(file_path, header=None, nrows=1, engine="openpyxl")
        val = raw.iloc[0, 0]
        if pd.notna(val):
            return str(val).strip() or None
    except Exception:
        pass
    return None


def load_mapping_files() -> dict[str, Any]:
    """
    Scan Excel spec files and index them by table name.

    Scanning strategy (applied in order):
      1. If both EXCEL_MASTER_PATH and EXCEL_MAPPING_PATH are set:
           - Only files inside EXCEL_MASTER_PATH  → Master Sheets
           - Only files inside EXCEL_MAPPING_PATH → Mapping Sheets
           - Files outside both folders are skipped
      2. If only EXCEL_MAPPING_PATH is set (no EXCEL_MASTER_PATH):
           - Files inside EXCEL_MAPPING_PATH → Mapping Sheets
           - All other files under EXCEL_DATA_PATH → Master Sheets
      3. If neither path is set but EXCEL_DATA_PATH is set:
           - All files under EXCEL_DATA_PATH → Master Sheets

    Indexes every file by its normalised table name (= file stem without extension).
    """
    global _loaded

    if not settings.EXCEL_DATA_PATH:
        _loaded = True
        return {"loaded": 0, "message": "EXCEL_DATA_PATH not configured"}

    base_path = Path(settings.EXCEL_DATA_PATH).expanduser().resolve()
    if not base_path.exists():
        _loaded = True
        return {"error": f"EXCEL_DATA_PATH does not exist: {base_path}"}

    # Resolve the optional master and mapping folders once
    master_root: Path | None = None
    if settings.EXCEL_MASTER_PATH:
        master_root = Path(settings.EXCEL_MASTER_PATH).expanduser().resolve()
        if not master_root.exists():
            log.warning("EXCEL_MASTER_PATH does not exist: %s — ignoring", master_root)
            master_root = None

    mapping_root: Path | None = None
    if settings.EXCEL_MAPPING_PATH:
        mapping_root = Path(settings.EXCEL_MAPPING_PATH).expanduser().resolve()
        if not mapping_root.exists():
            log.warning("EXCEL_MAPPING_PATH does not exist: %s — treating all files as master", mapping_root)
            mapping_root = None

    # Build the candidate file list
    explicit_roots = bool(master_root or mapping_root)
    _index.clear()
    loaded, skipped, errors = 0, 0, []

    excel_files: list[Path] = []
    for ext in ("*.xlsx", "*.xlsm", "*.xls"):
        excel_files.extend(sorted(base_path.rglob(ext)))

    for file_path in excel_files:
        # Determine type — skip files outside both explicit roots when both are configured
        if master_root and mapping_root:
            if file_path.is_relative_to(mapping_root):
                file_type = "mapping"
            elif file_path.is_relative_to(master_root):
                file_type = "master"
            else:
                log.debug("Skipping %s — outside both EXCEL_MASTER_PATH and EXCEL_MAPPING_PATH", file_path)
                skipped += 1
                continue
        else:
            file_type = _classify_file(file_path, mapping_root)

        header_arg = _header_arg(file_type)

        try:
            bq_table_name = _read_bq_table_name(file_path) if file_type == "mapping" else None

            sheets: dict[str, pd.DataFrame] = pd.read_excel(
                file_path, sheet_name=None, header=header_arg, engine="openpyxl"
            )
            rel_path = str(file_path.relative_to(base_path))
            table_name = file_path.stem  # filename (without extension) = table name
            norm_key = _normalize_name(table_name)

            parsed_sheets = {
                sheet_name: _parse_sheet(df, file_type)
                for sheet_name, df in sheets.items()
            }

            entry: dict[str, Any] = {
                "table_name": table_name,
                "normalized_name": norm_key,
                "file": rel_path,
                "type": file_type,
                "sheets": parsed_sheets,
            }
            if bq_table_name is not None:
                entry["bq_table_name"] = bq_table_name
            _index[norm_key] = entry
            loaded += 1
            log.info(
                "Indexed %s spec: %s → '%s' (%d sheet(s))",
                file_type, rel_path, table_name, len(parsed_sheets),
            )
        except Exception as exc:
            rel = str(file_path.relative_to(base_path)) if file_path.is_relative_to(base_path) else str(file_path)
            log.warning("Failed to parse spec file %s: %s", rel, exc)
            errors.append({"file": rel, "error": str(exc)})

    _loaded = True
    result: dict[str, Any] = {
        "loaded": loaded,
        "master": sum(1 for v in _index.values() if v["type"] == "master"),
        "mapping": sum(1 for v in _index.values() if v["type"] == "mapping"),
        "master_folder": str(master_root) if master_root else "not configured",
        "mapping_folder": str(mapping_root) if mapping_root else "not configured (all files treated as master)",
    }
    if skipped:
        result["skipped"] = skipped
    if errors:
        result["errors"] = errors
    return result


def _ensure_loaded() -> None:
    if not _loaded:
        load_mapping_files()


# ---------------------------------------------------------------------------
# Agent-facing tools
# ---------------------------------------------------------------------------

def list_mapping_tables() -> dict[str, Any]:
    """
    List all tables found in Excel Master and Mapping spec files.
    Returns table name, spec type, source file, available sheets,
    and (for mapping files) the count of source→target pairs.
    """
    _ensure_loaded()
    tables = []
    for v in _index.values():
        entry: dict[str, Any] = {
            "table_name": v["table_name"],
            "type": v["type"],
            "file": v["file"],
            "sheets": list(v["sheets"].keys()),
        }
        if v["type"] == "mapping":
            if "bq_table_name" in v:
                entry["bq_table_name"] = v["bq_table_name"]
            total_pairs = sum(
                s.get("mapped_columns", 0) for s in v["sheets"].values()
            )
            entry["total_source_target_pairs"] = total_pairs
        tables.append(entry)

    return {"tables": tables, "total": len(tables)}


def get_table_mapping(table_name: str) -> dict[str, Any]:
    """
    Retrieve the full Excel specification for a table.

    Returns:
      - file & type (master/mapping)
      - per-sheet column list, row count, source→target pairs, and sample rows
      - a 'citation' key for use in audit reports

    Args:
        table_name: the table name to look up (exact or approximate match)
    """
    _ensure_loaded()
    norm = _normalize_name(table_name)

    # 1. Exact match
    if norm in _index:
        entry = _index[norm]
        return {**entry, "citation": f"[{entry['file']} — {entry['type']} spec]"}

    # 2. Substring match (longest key that overlaps)
    candidates = [k for k in _index if norm in k or k in norm]
    if candidates:
        best = max(candidates, key=lambda k: len(set(norm.split("_")) & set(k.split("_"))))
        entry = _index[best]
        return {
            **entry,
            "citation": f"[{entry['file']} — {entry['type']} spec]",
            "note": f"Closest match for '{table_name}' → '{entry['table_name']}'",
        }

    return {
        "error": f"No spec file found matching '{table_name}'",
        "available_tables": [v["table_name"] for v in _index.values()],
        "hint": "Call list_mapping_tables() to see all available specs",
    }


# ---------------------------------------------------------------------------
# Schema Audit — MySQL source → BigQuery target reconciliation
# ---------------------------------------------------------------------------

import json
from datetime import datetime
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# MySQL → expected BigQuery type mapping
_MYSQL_TO_BQ: dict[str, str] = {
    "int": "INT64", "bigint": "INT64", "smallint": "INT64",
    "tinyint": "INT64", "mediumint": "INT64",
    "float": "FLOAT64", "double": "FLOAT64",
    "decimal": "NUMERIC",
    "varchar": "STRING", "char": "STRING", "text": "STRING",
    "longtext": "STRING", "mediumtext": "STRING", "tinytext": "STRING",
    "datetime": "DATETIME", "timestamp": "TIMESTAMP",
    "date": "DATE", "time": "TIME",
    "boolean": "BOOL", "bool": "BOOL", "bit": "BOOL",
    "json": "JSON",
    "blob": "BYTES", "longblob": "BYTES", "mediumblob": "BYTES",
}

# BQ Standard SQL aliases → canonical form
_BQ_ALIASES: dict[str, str] = {
    "INTEGER": "INT64", "INT": "INT64", "SMALLINT": "INT64",
    "BIGINT": "INT64", "BYTEINT": "INT64", "TINYINT": "INT64",
    "FLOAT": "FLOAT64",
    "DECIMAL": "NUMERIC", "BIGDECIMAL": "NUMERIC", "BIGNUMERIC": "NUMERIC",
    "BOOLEAN": "BOOL",
}

_AUDIT_COLUMNS = [
    "Column Name", "MySQL #", "BQ #",
    "MySQL Type", "Expected BQ Type", "Actual BQ Type",
    "BQ Description", "Status",
]

_STATUS_FILLS = {
    "🟢": PatternFill("solid", fgColor="C6EFCE"),
    "🟡": PatternFill("solid", fgColor="FFEB9C"),
    "🟠": PatternFill("solid", fgColor="FF7518"),
    "🔵": PatternFill("solid", fgColor="BDD7EE"),
}


def _mysql_to_bq(mysql_type: str) -> str:
    return _MYSQL_TO_BQ.get(mysql_type.strip().lower(), mysql_type.strip().upper())


def _canonical_bq(bq_type: str) -> str:
    return _BQ_ALIASES.get(bq_type.strip().upper(), bq_type.strip().upper())


def _audit_status(m: Optional[dict], b: Optional[dict], expected: str, actual: str) -> str:
    if m is None:
        return "🟠 BQ Only"
    if b is None:
        return "🔵 MySQL Only"
    if actual and expected and _canonical_bq(actual) != _canonical_bq(expected):
        return "🟡 Type Mismatch"
    return "🟢 Match"


def _safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\\\/\?\*\[\]:]", "_", name)[:31]
    if cleaned not in used:
        return cleaned
    for i in range(2, 1000):
        candidate = f"{cleaned[:31 - len(str(i)) - 1]}_{i}"
        if candidate not in used:
            return candidate
    return cleaned


def _fetch_mysql_metadata(client) -> "pd.DataFrame":
    sql = f"""
    SELECT
        h.table_name, h.eda_dataset_name, h.eda_view_name, h.deployed_to_prod,
        d.column_name, CAST(d.ordinal_position AS INT64) AS ordinal_position, d.data_type
    FROM `{settings.SCHEMA_HEADER_VIEW}` h
    JOIN `{settings.SCHEMA_DETAIL_VIEW}` d ON d.table_name = h.table_name
    WHERE h.is_streamed = 1
    ORDER BY h.table_name, d.ordinal_position
    """
    log.info("Querying MySQL metadata from BigQuery metadata views…")
    df = client.query(sql).to_dataframe()
    log.info("Retrieved %d column metadata rows.", len(df))
    return df


def _fetch_bq_schema(client, project: str, dataset: str, view: str) -> list[dict]:
    full_ref = f"{project}.{dataset}.{view}"
    try:
        table_ref = client.get_table(full_ref)
        return [
            {
                "column_name": field.name,
                "ordinal_position": idx,
                "data_type": field.field_type,
                "description": field.description or "",
            }
            for idx, field in enumerate(table_ref.schema, 1)
        ]
    except Exception as exc:
        log.warning("Could not fetch BQ schema for %s: %s", full_ref, exc)
        return []


def _reconcile(mysql_rows: list[dict], bq_rows: list[dict]) -> list[dict]:
    mysql_map = {r["column_name"]: r for r in mysql_rows}
    bq_map    = {r["column_name"]: r for r in bq_rows}
    mysql_ordered = sorted(mysql_map, key=lambda c: mysql_map[c]["ordinal_position"])
    bq_only = sorted(
        (c for c in bq_map if c not in mysql_map),
        key=lambda c: bq_map[c]["ordinal_position"],
    )
    rows = []
    for col in mysql_ordered + bq_only:
        m = mysql_map.get(col)
        b = bq_map.get(col)
        mysql_type  = m["data_type"].strip() if m else ""
        actual_bq   = b["data_type"].strip() if b else ""
        expected_bq = _mysql_to_bq(mysql_type) if mysql_type else ""
        rows.append({
            "Column Name":      col,
            "MySQL #":          int(m["ordinal_position"]) if m else "",
            "BQ #":             int(b["ordinal_position"]) if b else "",
            "MySQL Type":       mysql_type,
            "Expected BQ Type": expected_bq,
            "Actual BQ Type":   actual_bq,
            "BQ Description":   b["description"] if b else "",
            "Status":           _audit_status(m, b, expected_bq, actual_bq),
        })
    return rows


def _apply_audit_sheet_format(ws) -> None:
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    data_font   = Font(name="Courier New", size=11)
    thin = Side(style="thin", color="BFBFBF")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    status_col = None
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.value == "Status":
            status_col = cell.column
    col_widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            length = len(str(cell.value)) if cell.value is not None else 0
            col_widths[cell.column] = max(col_widths.get(cell.column, 0), length)
            if cell.row > 1:
                cell.font = data_font
                cell.border = cell_border
                cell.alignment = Alignment(vertical="center")
                if status_col and cell.column == status_col:
                    emoji = str(cell.value or "")[:2]
                    fill = _STATUS_FILLS.get(emoji)
                    if fill:
                        cell.fill = fill
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 3, 60)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 18


def _write_audit_excel(
    meta_df: "pd.DataFrame",
    tables: list[dict],
    bq_project: str,
    file_suffix: str,
    client,
    output_dir: str,
    timestamp: str,
) -> dict[str, Any]:
    from pathlib import Path as _Path
    output_file = str(_Path(output_dir) / f"schema_audit_{timestamp}_{file_suffix}.xlsx")
    all_rows: list[dict] = []
    table_dfs: dict[str, "pd.DataFrame"] = {}

    for tbl in tables:
        name    = tbl["table_name"]
        dataset = tbl["eda_dataset_name"]
        view    = tbl["eda_view_name"]
        log.info("[%s] → %s.%s.%s", name, bq_project, dataset, view)
        mysql_rows = (
            meta_df[meta_df["table_name"] == name][
                ["column_name", "ordinal_position", "data_type"]
            ].to_dict("records")
        )
        bq_rows = _fetch_bq_schema(client, bq_project, dataset, view)
        rows = _reconcile(mysql_rows, bq_rows)
        table_dfs[name] = pd.DataFrame(rows, columns=_AUDIT_COLUMNS)
        for row in rows:
            all_rows.append({"Table Name": name, **row})

    if not all_rows:
        return {"error": f"No columns reconciled for {file_suffix}"}

    summary_rows = [r for r in all_rows if not r["Status"].startswith("🟢")]
    summary_df = pd.DataFrame(
        summary_rows if summary_rows else [{}],
        columns=["Table Name"] + _AUDIT_COLUMNS,
    )
    used_names: set[str] = {"Summary"}
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        for tbl_name, df in table_dfs.items():
            sheet = _safe_sheet_name(tbl_name, used_names)
            used_names.add(sheet)
            df.to_excel(writer, sheet_name=sheet, index=False)

    wb = load_workbook(output_file)
    for sheet_name in wb.sheetnames:
        _apply_audit_sheet_format(wb[sheet_name])
    wb.save(output_file)

    total      = len(all_rows)
    matches    = sum(1 for r in all_rows if r["Status"].startswith("🟢"))
    mismatches = sum(1 for r in all_rows if r["Status"].startswith("🟡"))
    bq_only    = sum(1 for r in all_rows if r["Status"].startswith("🟠"))
    mysql_only = sum(1 for r in all_rows if r["Status"].startswith("🔵"))

    return {
        "output_file": output_file,
        "tables": len(table_dfs),
        "total_columns": total,
        "match": matches,
        "type_mismatch": mismatches,
        "bq_only": bq_only,
        "mysql_only": mysql_only,
    }


def _write_ddl_json(
    meta_df: "pd.DataFrame",
    tables: list[dict],
    file_suffix: str,
    output_dir: str,
    timestamp: str,
) -> str:
    from pathlib import Path as _Path
    output_file = str(_Path(output_dir) / f"schema_ddl_{timestamp}_{file_suffix}.json")
    ddl: list[dict] = []
    for tbl in tables:
        name = tbl["table_name"]
        cols = (
            meta_df[meta_df["table_name"] == name]
            .sort_values("ordinal_position")[["column_name", "data_type"]]
            .to_dict("records")
        )
        ddl.append({
            "Table": name,
            "Schema": [
                {"name": col["column_name"], "type": _mysql_to_bq(col["data_type"]), "mode": "NULLABLE"}
                for col in cols
            ],
        })
    with open(output_file, "w", encoding="utf-8") as fh:
        json.dump(ddl, fh, indent=2)
    return output_file


def run_schema_audit() -> dict[str, Any]:
    """
    Run the full MySQL → BigQuery schema reconciliation audit.

    Reads SCHEMA_METADATA_PROJECT, SCHEMA_BQ_PROJECT_PROD, SCHEMA_BQ_PROJECT_UAT,
    SCHEMA_HEADER_VIEW, SCHEMA_DETAIL_VIEW, and SCHEMA_AUDIT_OUTPUT_DIR from .env.

    Generates:
      - schema_audit_<timestamp>_prd.xlsx  (prod tables)
      - schema_audit_<timestamp>_uat.xlsx  (UAT tables)
      - schema_ddl_<timestamp>_prd.json
      - schema_ddl_<timestamp>_uat.json

    Returns a summary dict with file paths and column-level stats.
    """
    if not settings.SCHEMA_METADATA_PROJECT:
        return {"error": "SCHEMA_METADATA_PROJECT is not set in .env"}
    if not settings.SCHEMA_HEADER_VIEW or not settings.SCHEMA_DETAIL_VIEW:
        return {"error": "SCHEMA_HEADER_VIEW and SCHEMA_DETAIL_VIEW must be set in .env"}

    from google.cloud import bigquery
    client = bigquery.Client(project=settings.SCHEMA_METADATA_PROJECT)

    meta_df = _fetch_mysql_metadata(client)
    if meta_df.empty:
        return {"error": "No streamed tables found in HEADER_VIEW"}

    tbl_cols = ["table_name", "eda_dataset_name", "eda_view_name", "deployed_to_prod"]
    all_tables  = meta_df[tbl_cols].drop_duplicates().to_dict("records")
    prod_tables = [t for t in all_tables if t.get("deployed_to_prod") == 1]
    uat_tables  = [t for t in all_tables if t.get("deployed_to_prod") != 1]

    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = settings.SCHEMA_AUDIT_OUTPUT_DIR
    results: dict[str, Any] = {
        "tables_found": len(all_tables),
        "prod_tables": len(prod_tables),
        "uat_tables": len(uat_tables),
    }

    if prod_tables and settings.SCHEMA_BQ_PROJECT_PROD:
        results["prod"] = _write_audit_excel(
            meta_df, prod_tables, settings.SCHEMA_BQ_PROJECT_PROD, "prd", client, output_dir, timestamp,
        )
        results["prod"]["ddl_json"] = _write_ddl_json(meta_df, prod_tables, "prd", output_dir, timestamp)

    if uat_tables:
        uat_project = settings.SCHEMA_BQ_PROJECT_UAT or settings.SCHEMA_METADATA_PROJECT
        results["uat"] = _write_audit_excel(
            meta_df, uat_tables, uat_project, "uat", client, output_dir, timestamp,
        )
        results["uat"]["ddl_json"] = _write_ddl_json(meta_df, uat_tables, "uat", output_dir, timestamp)

    return results
